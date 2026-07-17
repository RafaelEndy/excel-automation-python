"""
Automação de Relatório de Vendas - le uma planilha bruta (com formatos
inconsistentes, duplicatas e valores faltando), limpa os dados e gera
um relatório Excel formatado com resumo, grafico e dados tratados.

Uso: python src/gerar_relatorio.py
Entrada:  data/vendas_brutas.csv
Saida:    output/relatorio_vendas.xlsx
          output/grafico_receita_categoria.png
"""

import os

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_PATH = os.path.join(BASE_DIR, "data", "vendas_brutas.csv")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "relatorio_vendas.xlsx")
OUTPUT_CHART_PNG = os.path.join(OUTPUT_DIR, "grafico_receita_categoria.png")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


def load_raw_data(path: str) -> pd.DataFrame:
    """Le o CSV bruto (separado por ';', formatos de data variados)."""
    return pd.read_csv(path, sep=";", dtype=str)


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza os dados brutos:
      - remove espacos extras de texto
      - padroniza categorias para Title Case
      - converte datas em formatos diferentes para um unico padrao
      - converte colunas numericas, tratando valores ausentes
      - remove linhas duplicadas
    """
    df = df.copy()

    for col in ["Produto", "Categoria"]:
        df[col] = df[col].str.strip()

    df["Categoria"] = df["Categoria"].str.title()

    df["Data"] = pd.to_datetime(df["Data"].str.strip(), format="mixed", dayfirst=True)

    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
    df["Valor Unitario"] = pd.to_numeric(df["Valor Unitario"], errors="coerce")

    # Preenche quantidade/valor ausentes com a mediana do mesmo produto;
    # se o produto so aparecer uma vez sem valor, usa a mediana geral como fallback
    for col in ["Quantidade", "Valor Unitario"]:
        df[col] = df.groupby("Produto")[col].transform(lambda s: s.fillna(s.median()))
        df[col] = df[col].fillna(df[col].median())

    df = df.drop_duplicates()

    df["Receita"] = df["Quantidade"] * df["Valor Unitario"]

    return df.sort_values("Data").reset_index(drop=True)


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega a receita total por categoria, ordenada da maior para a menor."""
    summary = (
        df.groupby("Categoria")["Receita"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )
    summary["Receita"] = summary["Receita"].round(2)
    return summary


def generate_chart_image(summary: pd.DataFrame, path: str) -> None:
    """Gera um grafico de barras (receita por categoria) como imagem PNG."""
    plt.figure(figsize=(7, 4))
    plt.bar(summary["Categoria"], summary["Receita"], color="#1F4E78")
    plt.title("Receita por Categoria")
    plt.ylabel("Receita (R$)")
    plt.xticks(rotation=15)
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def style_header_row(worksheet, row: int, num_columns: int) -> None:
    """Aplica estilo (fundo azul, fonte branca em negrito) na linha de cabecalho."""
    for col in range(1, num_columns + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(worksheet, df: pd.DataFrame, start_col: int = 1) -> None:
    """Ajusta a largura das colunas com base no conteudo mais longo."""
    for i, col in enumerate(df.columns, start=start_col):
        max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = max_len


def write_excel_report(df: pd.DataFrame, summary: pd.DataFrame, path: str) -> None:
    """Grava o relatorio final em Excel: aba de Resumo e aba de Dados Tratados."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Resumo", index=False, startrow=2)
        df.to_excel(writer, sheet_name="Dados Tratados", index=False)

        wb = writer.book

        # --- Aba Resumo ---
        ws_summary = writer.sheets["Resumo"]
        ws_summary["A1"] = "Relatorio de Vendas — Resumo por Categoria"
        ws_summary["A1"].font = TITLE_FONT
        style_header_row(ws_summary, row=3, num_columns=len(summary.columns))
        autofit_columns(ws_summary, summary)

        chart = BarChart()
        chart.title = "Receita por Categoria"
        chart.y_axis.title = "Receita (R$)"
        data_ref = Reference(
            ws_summary, min_col=2, min_row=3, max_row=3 + len(summary)
        )
        cats_ref = Reference(
            ws_summary, min_col=1, min_row=4, max_row=3 + len(summary)
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_summary.add_chart(chart, f"A{6 + len(summary)}")

        # --- Aba Dados Tratados ---
        ws_data = writer.sheets["Dados Tratados"]
        style_header_row(ws_data, row=1, num_columns=len(df.columns))
        autofit_columns(ws_data, df)


def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Lendo dados brutos...")
    raw_df = load_raw_data(INPUT_PATH)
    print(f"  {len(raw_df)} linhas lidas de '{INPUT_PATH}'.")

    print("Limpando e tratando dados...")
    clean_df = clean_data(raw_df)
    print(f"  {len(clean_df)} linhas apos limpeza (duplicatas removidas).")

    print("Calculando resumo por categoria...")
    summary = build_summary(clean_df)

    print("Gerando grafico...")
    generate_chart_image(summary, OUTPUT_CHART_PNG)

    print("Gravando relatorio Excel...")
    write_excel_report(clean_df, summary, OUTPUT_XLSX)

    print(f"\nRelatorio gerado com sucesso: {OUTPUT_XLSX}")
    print(f"Grafico gerado com sucesso:   {OUTPUT_CHART_PNG}")
    print("\nResumo por categoria:")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
